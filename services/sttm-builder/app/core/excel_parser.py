"""Excel/CSV mapping file parser.

Parses BBI S-T Mapping Excel format into structured column mappings.
Handles .xlsx, .xls, and .csv files.
"""
from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

EXPECTED_HEADERS = {
    "target field name", "source field names", "target table name",
    "source dataset name", "pre processing rules", "field definition",
    "target field data type", "field type",
}


@dataclass
class ExcelColumnMapping:
    target_table: str
    target_field: str
    target_data_type: str
    source_field: str
    source_dataset: str
    preprocessing_rule: str | None = None
    field_definition: str | None = None
    field_type: str | None = None
    processing_order: int | None = None
    depends_on: str | None = None
    contains_pii: bool = False

    def to_dict(self) -> dict[str, Any]:
        return {
            "target_table": self.target_table,
            "target_field": self.target_field,
            "target_data_type": self.target_data_type,
            "source_field": self.source_field,
            "source_dataset": self.source_dataset,
            "preprocessing_rule": self.preprocessing_rule,
            "field_definition": self.field_definition,
            "field_type": self.field_type,
            "processing_order": self.processing_order,
            "depends_on": self.depends_on,
            "contains_pii": self.contains_pii,
        }


@dataclass
class ParsedExcelMapping:
    target_tables: list[str] = field(default_factory=list)
    source_datasets: list[str] = field(default_factory=list)
    column_mappings: list[ExcelColumnMapping] = field(default_factory=list)
    sheet_name: str = ""
    total_rows: int = 0
    parse_warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "target_tables": self.target_tables,
            "source_datasets": self.source_datasets,
            "column_mappings": [m.to_dict() for m in self.column_mappings],
            "sheet_name": self.sheet_name,
            "total_rows": self.total_rows,
            "parse_warnings": self.parse_warnings,
            "stats": {
                "target_tables": len(self.target_tables),
                "source_datasets": len(self.source_datasets),
                "columns_mapped": len(self.column_mappings),
                "columns_with_rules": sum(
                    1 for m in self.column_mappings if m.preprocessing_rule
                ),
            },
        }


def parse_excel_mapping(content: bytes, filename: str) -> ParsedExcelMapping:
    """Parse an Excel or CSV file into structured mapping data."""
    if filename.lower().endswith(".csv"):
        return _parse_csv(content)
    if not HAS_OPENPYXL:
        result = ParsedExcelMapping()
        result.parse_warnings.append("openpyxl not installed — cannot parse Excel files")
        return result
    return _parse_xlsx(content)


def _parse_xlsx(content: bytes) -> ParsedExcelMapping:
    """Parse .xlsx file using openpyxl."""
    result = ParsedExcelMapping()

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Find the best sheet (prefer "S-T Mapping Rules" or first sheet with mapping headers)
    target_sheet = None
    for name in wb.sheetnames:
        if "mapping" in name.lower() and "rule" in name.lower():
            target_sheet = name
            break
    if not target_sheet:
        for name in wb.sheetnames:
            if "s-t" in name.lower() or "mapping" in name.lower():
                target_sheet = name
                break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    result.sheet_name = target_sheet
    ws = wb[target_sheet]

    # Read header row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result.parse_warnings.append("Empty sheet")
        return result

    headers = [str(h).strip().lower().replace("\n", " ") if h else "" for h in rows[0]]

    # Map header indices
    col_idx = {}
    for i, h in enumerate(headers):
        if "target table" in h:
            col_idx["target_table"] = i
        elif "target field name" in h:
            col_idx["target_field"] = i
        elif "target field data type" in h:
            col_idx["target_data_type"] = i
        elif "source field" in h:
            col_idx["source_field"] = i
        elif "source dataset" in h:
            col_idx["source_dataset"] = i
        elif "pre processing" in h or "preprocessing" in h:
            col_idx["preprocessing"] = i
        elif "field definition" in h:
            col_idx["definition"] = i
        elif "field type" in h and "data" not in h:
            col_idx["field_type"] = i
        elif "processing order" in h:
            col_idx["order"] = i
        elif "depends on" in h or "field depends" in h:
            col_idx["depends_on"] = i
        elif "pii" in h:
            col_idx["pii"] = i

    if "target_field" not in col_idx:
        result.parse_warnings.append(f"Could not find 'Target Field Name' column in headers: {headers[:10]}")
        return result

    # Parse data rows
    target_tables = set()
    source_datasets = set()

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue

        def _get(key: str) -> str | None:
            idx = col_idx.get(key)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        target_field = _get("target_field")
        if not target_field:
            continue

        target_table = _get("target_table") or ""
        source_field = _get("source_field") or ""
        source_dataset = _get("source_dataset") or ""
        target_data_type = _get("target_data_type") or ""
        preprocessing = _get("preprocessing")
        definition = _get("definition")
        field_type = _get("field_type")
        order_str = _get("order")
        depends_on = _get("depends_on")
        pii_val = _get("pii")

        if target_table:
            target_tables.add(target_table)
        if source_dataset:
            source_datasets.add(source_dataset)

        mapping = ExcelColumnMapping(
            target_table=target_table,
            target_field=target_field,
            target_data_type=target_data_type,
            source_field=source_field,
            source_dataset=source_dataset,
            preprocessing_rule=preprocessing,
            field_definition=definition,
            field_type=field_type,
            processing_order=int(order_str) if order_str and order_str.isdigit() else None,
            depends_on=depends_on,
            contains_pii=str(pii_val).lower() in ("true", "1", "yes") if pii_val else False,
        )
        result.column_mappings.append(mapping)

    result.target_tables = sorted(target_tables)
    result.source_datasets = sorted(source_datasets)
    result.total_rows = len(result.column_mappings)

    return result


def _parse_csv(content: bytes) -> ParsedExcelMapping:
    """Parse CSV mapping file."""
    result = ParsedExcelMapping()
    result.sheet_name = "csv"

    text = content.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    target_tables = set()
    source_datasets = set()

    for row in reader:
        target_field = None
        source_field = None
        target_table = None

        for key, val in row.items():
            key_lower = key.lower().strip()
            if "target field name" in key_lower:
                target_field = val
            elif "source field" in key_lower:
                source_field = val
            elif "target table" in key_lower:
                target_table = val

        if target_field:
            mapping = ExcelColumnMapping(
                target_table=target_table or "",
                target_field=target_field,
                target_data_type=row.get("Target Field Data Type", ""),
                source_field=source_field or "",
                source_dataset=row.get("Source Dataset Name", ""),
                preprocessing_rule=row.get("Pre Processing Rules"),
                field_definition=row.get("Field Definition"),
            )
            result.column_mappings.append(mapping)
            if target_table:
                target_tables.add(target_table)

    result.target_tables = sorted(target_tables)
    result.source_datasets = sorted(source_datasets)
    result.total_rows = len(result.column_mappings)
    return result
