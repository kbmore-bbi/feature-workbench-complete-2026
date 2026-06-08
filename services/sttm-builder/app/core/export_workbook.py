from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re
from typing import Any

from snowflake.snowpark import Session

from app.core.exceptions import SnowflakeQueryError
from app.schema.export_workbook import WorkbookExportRequest
from app.schema.mapping_sql import MappingSqlPreviewRow


class WorkbookExportService:
    def __init__(self, *, session: Session) -> None:
        self._session = session
        self._cell_alignment = None

    def build_workbook(self, body: WorkbookExportRequest) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - environment guard
            raise RuntimeError(
                "Excel export dependencies are not installed. Install 'openpyxl' in services/sttm-builder/.venv to enable workbook downloads."
            ) from exc
        try:
            from PIL import Image as PilImage
            from PIL import ImageDraw, ImageFont
        except ImportError:  # pragma: no cover - optional visual enhancement
            PilImage = None
            ImageDraw = None
            ImageFont = None

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        section_fill = PatternFill("solid", fgColor="EEF6FF")
        source_fill = PatternFill("solid", fgColor="E8F4FD")
        derived_fill = PatternFill("solid", fgColor="F3E8FF")
        target_fill = PatternFill("solid", fgColor="DCFCE7")
        edge_fill = PatternFill("solid", fgColor="F8FAFC")
        self._cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        base_font = Font(name="Arial", size=10)
        header_font = Font(name="Arial", size=10, bold=True, color="1F2937")
        title_font = Font(name="Arial", size=13, bold=True, color="111827")

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        derived_sheet = workbook.create_sheet("Derived Sources")
        sttm_sheet = workbook.create_sheet("STTM")
        sql_sheet = workbook.create_sheet("Generated SQL")
        table_lineage_sheet = workbook.create_sheet("Table Lineage")
        column_lineage_sheet = workbook.create_sheet("Column Lineage")
        dbt_sheet = workbook.create_sheet("DBT Code")

        preview_warning: str | None = None
        try:
            preview_rows = self._preview_rows(body.preview_sql)
            preview_values = [row.values for row in preview_rows]
        except Exception as exc:  # pragma: no cover - export should fail open here
            preview_warning = self._summarize_preview_error(exc)
            preview_rows = []
            preview_values = []

        self._build_summary_sheet(
            summary_sheet,
            body,
            title_font,
            header_font,
            base_font,
            preview_warning=preview_warning,
        )
        self._build_derived_sheet(
            derived_sheet,
            body,
            get_column_letter,
            title_font,
            header_font,
            base_font,
            section_fill,
            header_fill,
            thin_border,
        )
        self._build_sttm_sheet(
            sttm_sheet,
            body,
            preview_values,
            get_column_letter,
            title_font,
            header_font,
            base_font,
            section_fill,
            header_fill,
            thin_border,
        )
        self._build_sql_sheet(
            sql_sheet,
            body,
            title_font,
            base_font,
            section_fill,
            thin_border,
        )
        self._build_mermaid_sheet(
            table_lineage_sheet,
            title="Table-level lineage",
            mermaid_text=body.lineage_table_mermaid,
            fallback_message="No table-level lineage diagram is available yet.",
            openpyxl_image_cls=OpenpyxlImage,
            pil_image_cls=PilImage,
            image_draw_cls=ImageDraw,
            image_font_cls=ImageFont,
            title_font=title_font,
            header_font=header_font,
            base_font=base_font,
            section_fill=section_fill,
            thin_border=thin_border,
            source_fill=source_fill,
            derived_fill=derived_fill,
            target_fill=target_fill,
            edge_fill=edge_fill,
        )
        self._build_mermaid_sheet(
            column_lineage_sheet,
            title="Column-level lineage",
            mermaid_text=body.lineage_column_mermaid,
            fallback_message="No column-level lineage diagram is available yet.",
            openpyxl_image_cls=OpenpyxlImage,
            pil_image_cls=PilImage,
            image_draw_cls=ImageDraw,
            image_font_cls=ImageFont,
            title_font=title_font,
            header_font=header_font,
            base_font=base_font,
            section_fill=section_fill,
            thin_border=thin_border,
            source_fill=source_fill,
            derived_fill=derived_fill,
            target_fill=target_fill,
            edge_fill=edge_fill,
        )
        self._build_dbt_sheet(
            dbt_sheet,
            body,
            title_font,
            header_font,
            base_font,
            section_fill,
            thin_border,
        )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _preview_rows(self, preview_sql: str) -> list[MappingSqlPreviewRow]:
        sql_text = (preview_sql or "").strip().rstrip(";")
        if not sql_text or sql_text.startswith("--"):
            return []
        try:
            dataframe = self._session.sql(
                "SELECT * FROM (\n"
                f"{sql_text}\n"
                ") AS EXPORT_PREVIEW LIMIT 2"
            )
            return [
                MappingSqlPreviewRow(values=row.as_dict(recursive=True))
                for row in dataframe.collect()
            ]
        except Exception as exc:
            raise SnowflakeQueryError(
                f"Failed to load preview sample rows for the workbook export: {exc}"
            ) from exc

    def _build_summary_sheet(
        self,
        sheet,
        body: WorkbookExportRequest,
        title_font,
        header_font,
        base_font,
        *,
        preview_warning: str | None = None,
    ) -> None:
        rows = [
            ("STTM Export Summary", None),
            ("Project", body.project_name or "STTM Export"),
            ("Created by", body.created_by or "Unknown"),
            ("Generated at", body.created_at or datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Version details", body.version_label or "Current builder session"),
            ("Target table", body.target_table.qualified_name if body.target_table else "Not selected"),
            ("Source tables", ", ".join(table.qualified_name for table in body.source_tables) or "None"),
            (
                "DBT export status",
                body.dbt_conversion.status if body.dbt_conversion else "DBT conversion not attached",
            ),
            (
                "Sample preview rows",
                "Unavailable" if preview_warning else "Included when preview SQL returned rows",
            ),
            ("SQL export variant", body.sql_variant_label or "Current builder SQL"),
            ("Overview", body.summary_narrative or "Summary narrative not available."),
        ]
        for idx, (label, value) in enumerate(rows, start=1):
            cell_a = sheet.cell(row=idx, column=1, value=label)
            cell_b = sheet.cell(row=idx, column=2, value=value)
            cell_a.font = title_font if idx == 1 else header_font
            cell_b.font = base_font
        if preview_warning:
            warning_row = len(rows) + 1
            sheet.cell(row=warning_row, column=1, value="Preview warning").font = header_font
            sheet.cell(row=warning_row, column=2, value=preview_warning).font = base_font
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 96

    def _build_derived_sheet(self, sheet, body: WorkbookExportRequest, get_column_letter, title_font, header_font, base_font, section_fill, header_fill, thin_border) -> None:
        row = 1
        row = self._write_section_title(sheet, row, "Derived sources", title_font, section_fill)
        derived_headers = [
            "Derived Source Name",
            "Source Tables",
            "Base Source Tables",
            "Semantic View",
            "Semantic Bundle",
            "SQL",
        ]
        row = self._write_table_header(sheet, row, derived_headers, header_font, header_fill, thin_border)
        derived_items = body.derived_sources or []
        if derived_items:
            for item in derived_items:
                values = [
                    item.derived_source_name,
                    ", ".join(table.qualified_name for table in item.source_tables),
                    ", ".join(table.qualified_name for table in item.base_source_tables),
                    item.semantic_view_name or "",
                    item.semantic_bundle_label or "",
                    item.sql_text or "",
                ]
                self._write_row(sheet, row, values, base_font, thin_border)
                row += 1
        else:
            self._write_row(sheet, row, ["No derived sources selected.", "", "", "", "", ""], base_font, thin_border)
            row += 1

        row += 1
        row = self._write_section_title(sheet, row, "Source relationships", title_font, section_fill)
        relationship_headers = [
            "Left Table",
            "Join Type",
            "Right Table",
            "Join Conditions",
            "Constraint Name",
            "Source",
        ]
        row = self._write_table_header(sheet, row, relationship_headers, header_font, header_fill, thin_border)
        if body.relationships:
            for relationship in body.relationships:
                conditions = ", ".join(
                    f"{condition.left_column} {condition.operator} {condition.right_column}"
                    for condition in relationship.conditions
                )
                self._write_row(
                    sheet,
                    row,
                    [
                        relationship.left_table.qualified_name,
                        relationship.join_type,
                        relationship.right_table.qualified_name,
                        conditions,
                        relationship.constraint_name or "",
                        relationship.source or "",
                    ],
                    base_font,
                    thin_border,
                )
                row += 1
        else:
            self._write_row(sheet, row, ["No join relationships available.", "", "", "", "", ""], base_font, thin_border)
            row += 1

        for idx, width in enumerate([34, 12, 34, 48, 20, 16], start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    def _build_sttm_sheet(self, sheet, body: WorkbookExportRequest, preview_values: list[dict[str, Any]], get_column_letter, title_font, header_font, base_font, section_fill, header_fill, thin_border) -> None:
        row = 1
        row = self._write_section_title(sheet, row, "List of different sources join conditions", title_font, section_fill)
        row = self._write_table_header(
            sheet,
            row,
            ["Source Table", "Target Table", "Join Type", "Join Condition"],
            header_font,
            header_fill,
            thin_border,
        )
        if body.relationships:
            for relationship in body.relationships:
                self._write_row(
                    sheet,
                    row,
                    [
                        relationship.left_table.qualified_name,
                        relationship.right_table.qualified_name,
                        relationship.join_type,
                        ", ".join(
                            f"{condition.left_column} {condition.operator} {condition.right_column}"
                            for condition in relationship.conditions
                        ),
                    ],
                    base_font,
                    thin_border,
                )
                row += 1
        else:
            self._write_row(sheet, row, ["No joins configured.", "", "", ""], base_font, thin_border)
            row += 1

        row += 1
        row = self._write_section_title(sheet, row, "Filters that should be applied", title_font, section_fill)
        row = self._write_table_header(sheet, row, ["Filters"], header_font, header_fill, thin_border)
        self._write_row(sheet, row, [body.filters_sql or "No filters configured."], base_font, thin_border)
        row += 2

        row = self._write_section_title(sheet, row, "Target mappings", title_font, section_fill)
        headers = [
            "id",
            "Partner",
            "Schema",
            "Target Table Name",
            "Target Field Name",
            "Target Field Data Type",
            "pk?",
            "contains_pii",
            "Field Type",
            "Field Depends On",
            "Processing Order",
            "Pre Processing Rules",
            "Source Field Names",
            "Source Dataset Name",
            "Field Definition",
            "Example 1",
            "Example 2",
        ]
        row = self._write_table_header(sheet, row, headers, header_font, header_fill, thin_border)
        target_schema = body.target_table.schema if body.target_table else ""
        target_table_name = body.target_table.table if body.target_table else ""
        mappings = [mapping for mapping in body.mappings if (mapping.status or "").upper() == "MAPPED"]
        for index, mapping in enumerate(mappings, start=1):
            source_columns = mapping.source_columns or []
            if not source_columns and mapping.source_column:
                source_columns = [part.strip() for part in mapping.source_column.split(",") if part.strip()]
            if not source_columns:
                source_columns = [""]
            start_row = row
            example_1 = preview_values[0].get(mapping.target_column) if len(preview_values) > 0 else ""
            example_2 = preview_values[1].get(mapping.target_column) if len(preview_values) > 1 else ""
            for source_idx, source_column in enumerate(source_columns):
                dataset_name = ".".join(source_column.split(".")[:-1]) if "." in source_column else ""
                values = [
                    index if source_idx == 0 else "",
                    body.project_name or "",
                    target_schema if source_idx == 0 else "",
                    target_table_name if source_idx == 0 else "",
                    mapping.target_column if source_idx == 0 else "",
                    mapping.target_type if source_idx == 0 else "",
                    "No" if source_idx == 0 else "",
                    "" if source_idx == 0 else "",
                    ("TRANSFORMED" if mapping.expression or (mapping.rule and mapping.rule not in {"Direct", "Select..."}) else "RAW") if source_idx == 0 else "",
                    ", ".join(source_columns) if source_idx == 0 else "",
                    mapping.nl_rule or "" if source_idx == 0 else "",
                    (mapping.expression or mapping.rule or "") if source_idx == 0 else "",
                    source_column,
                    dataset_name,
                    mapping.description or "" if source_idx == 0 else "",
                    example_1 if source_idx == 0 else "",
                    example_2 if source_idx == 0 else "",
                ]
                self._write_row(sheet, row, values, base_font, thin_border)
                row += 1
            if row - start_row > 1:
                for column in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 15, 16, 17):
                    sheet.merge_cells(start_row=start_row, start_column=column, end_row=row - 1, end_column=column)

        widths = [8, 20, 18, 24, 24, 18, 8, 12, 14, 22, 18, 24, 26, 26, 32, 16, 16]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    def _build_sql_sheet(
        self,
        sheet,
        body: WorkbookExportRequest,
        title_font,
        base_font,
        section_fill,
        thin_border,
    ) -> None:
        row = 1
        row = self._write_section_title(
            sheet,
            row,
            body.sql_variant_label or "Final SQL used for export",
            title_font,
            section_fill,
        )
        for line in (body.generated_sql or "No generated SQL available.").splitlines() or [""]:
            self._write_row(sheet, row, [line], base_font, thin_border)
            row += 1
        row += 1
        row = self._write_section_title(sheet, row, "Source query foundation", title_font, section_fill)
        for line in (body.source_query_sql or "No source query SQL available.").splitlines() or [""]:
            self._write_row(sheet, row, [line], base_font, thin_border)
            row += 1
        sheet.column_dimensions["A"].width = 120

    def _build_mermaid_sheet(
        self,
        sheet,
        *,
        title: str,
        mermaid_text: str | None,
        fallback_message: str,
        openpyxl_image_cls,
        pil_image_cls,
        image_draw_cls,
        image_font_cls,
        title_font,
        header_font,
        base_font,
        section_fill,
        thin_border,
        source_fill,
        derived_fill,
        target_fill,
        edge_fill,
    ) -> None:
        row = 1
        row = self._write_section_title(sheet, row, title, title_font, section_fill)
        nodes, edges = self._parse_mermaid_flowchart(mermaid_text)
        if not edges:
            for line in (fallback_message if not mermaid_text else mermaid_text).splitlines() or [""]:
                self._write_row(sheet, row, [line], base_font, thin_border)
                row += 1
            sheet.column_dimensions["A"].width = 120
            return

        row = self._write_table_header(
            sheet,
            row,
            ["Legend", "Meaning"],
            header_font,
            section_fill,
            thin_border,
        )
        legend_rows = [
            ("Source", source_fill),
            ("Derived", derived_fill),
            ("Target", target_fill),
        ]
        for label, fill in legend_rows:
            self._write_row(sheet, row, [label, f"{label} node"], base_font, thin_border)
            sheet.cell(row=row, column=1).fill = fill
            row += 1

        diagram_image = self._render_lineage_diagram_image(
            nodes,
            edges,
            pil_image_cls=pil_image_cls,
            image_draw_cls=image_draw_cls,
            image_font_cls=image_font_cls,
        )
        if diagram_image is not None and openpyxl_image_cls is not None:
            sheet.add_image(openpyxl_image_cls(diagram_image), "A6")
            row = max(row, 30)

        row = self._write_table_header(
            sheet,
            row,
            ["From", "", "Flow", "", "To", "", "Meaning"],
            header_font,
            section_fill,
            thin_border,
        )
        for edge in edges:
            from_node = nodes.get(edge["from"], {"id": edge["from"], "label": edge["from"]})
            to_node = nodes.get(edge["to"], {"id": edge["to"], "label": edge["to"]})
            row = self._write_lineage_edge_row(
                sheet,
                row,
                from_node=from_node,
                to_node=to_node,
                edge_label=edge["label"],
                edge_kind=edge["kind"],
                base_font=base_font,
                thin_border=thin_border,
                source_fill=source_fill,
                derived_fill=derived_fill,
                target_fill=target_fill,
                edge_fill=edge_fill,
            )

        for column, width in {
            "A": 28,
            "B": 4,
            "C": 14,
            "D": 14,
            "E": 28,
            "F": 4,
            "G": 46,
        }.items():
            sheet.column_dimensions[column].width = width

    def _build_dbt_sheet(
        self,
        sheet,
        body: WorkbookExportRequest,
        title_font,
        header_font,
        base_font,
        section_fill,
        thin_border,
    ) -> None:
        row = 1
        row = self._write_section_title(sheet, row, "DBT conversion summary", title_font, section_fill)
        summary_rows = [
            ("Status", body.dbt_conversion.status if body.dbt_conversion else "Not generated yet"),
            ("Action", body.dbt_conversion.action if body.dbt_conversion else ""),
            ("Materialization", body.dbt_conversion.materialization if body.dbt_conversion else ""),
            (
                "Summary",
                body.dbt_conversion.message
                if body.dbt_conversion and body.dbt_conversion.message
                else "DBT conversion was not ready when this workbook was downloaded.",
            ),
            (
                "Materialization reason",
                body.dbt_conversion.materialization_reason if body.dbt_conversion else "",
            ),
        ]
        for label, value in summary_rows:
            self._write_row(sheet, row, [label, value], base_font, thin_border)
            sheet.cell(row=row, column=1).font = header_font
            row += 1

        row += 1
        row = self._write_section_title(sheet, row, "Generated files", title_font, section_fill)
        if not body.dbt_conversion or not body.dbt_conversion.generated_files:
            self._write_row(
                sheet,
                row,
                ["No DBT model files were attached to this export yet."],
                base_font,
                thin_border,
            )
            row += 1
        else:
            for file in body.dbt_conversion.generated_files:
                row = self._write_code_file(sheet, row, file.file_path, file.content, title_font, base_font, section_fill, thin_border)

        row += 1
        row = self._write_section_title(sheet, row, "Schema files", title_font, section_fill)
        if not body.dbt_conversion or not body.dbt_conversion.schema_files:
            self._write_row(sheet, row, ["No schema files were attached."], base_font, thin_border)
            row += 1
        else:
            for file in body.dbt_conversion.schema_files:
                row = self._write_code_file(sheet, row, file.file_path, file.content, title_font, base_font, section_fill, thin_border)

        row += 1
        row = self._write_section_title(sheet, row, "Sources YAML update", title_font, section_fill)
        if not body.dbt_conversion or not body.dbt_conversion.source_update:
            self._write_row(sheet, row, ["No source YAML update was attached."], base_font, thin_border)
            row += 1
        else:
            source_update = body.dbt_conversion.source_update
            row = self._write_code_file(
                sheet,
                row,
                f"{source_update.file_path} ({source_update.action})",
                source_update.content or "",
                title_font,
                base_font,
                section_fill,
                thin_border,
            )

        sheet.column_dimensions["A"].width = 42
        sheet.column_dimensions["B"].width = 120

    def _write_code_file(
        self,
        sheet,
        row: int,
        file_path: str,
        content: str,
        title_font,
        base_font,
        section_fill,
        thin_border,
    ) -> int:
        row = self._write_section_title(sheet, row, file_path, title_font, section_fill)
        for line in (content or "").splitlines() or [""]:
            self._write_row(sheet, row, [line], base_font, thin_border)
            row += 1
        return row + 1

    def _write_section_title(self, sheet, row: int, title: str, title_font, section_fill) -> int:
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = title_font
        cell.fill = section_fill
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        return row + 1

    def _write_table_header(self, sheet, row: int, headers: list[str], header_font, header_fill, thin_border) -> int:
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=index, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        return row + 1

    def _write_row(self, sheet, row: int, values: list[Any], base_font, thin_border) -> None:
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.font = base_font
            cell.border = thin_border
            if self._cell_alignment is not None:
                cell.alignment = self._cell_alignment

    def _write_lineage_edge_row(
        self,
        sheet,
        row: int,
        *,
        from_node: dict[str, str],
        to_node: dict[str, str],
        edge_label: str | None,
        edge_kind: str,
        base_font,
        thin_border,
        source_fill,
        derived_fill,
        target_fill,
        edge_fill,
    ) -> int:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

        from_cell = sheet.cell(row=row, column=1, value=from_node["label"])
        flow_cell = sheet.cell(row=row, column=3, value="-->" if edge_kind == "solid" else "-.-")
        to_cell = sheet.cell(row=row, column=5, value=to_node["label"])
        meaning_cell = sheet.cell(row=row, column=7, value=edge_label or "maps to")

        for cell in (from_cell, flow_cell, to_cell, meaning_cell):
            cell.font = base_font
            cell.border = thin_border
            if self._cell_alignment is not None:
                cell.alignment = self._cell_alignment

        from_cell.fill = self._node_fill_for_id(from_node["id"], source_fill, derived_fill, target_fill)
        to_cell.fill = self._node_fill_for_id(to_node["id"], source_fill, derived_fill, target_fill)
        flow_cell.fill = edge_fill

        return row + 1

    @staticmethod
    def _node_fill_for_id(node_id: str, source_fill, derived_fill, target_fill):
        if node_id.startswith("drv_"):
            return derived_fill
        if node_id.startswith("target_col_") or node_id.startswith("target_tbl_"):
            return target_fill
        if node_id.startswith("source_col_") or node_id.startswith("source_tbl_"):
            return source_fill
        if node_id.startswith("tbl_"):
            return target_fill if "_DW_" in node_id or "_target_" in node_id.lower() else source_fill
        return source_fill

    @staticmethod
    def _parse_mermaid_flowchart(mermaid_text: str | None) -> tuple[dict[str, dict[str, str]], list[dict[str, str | None]]]:
        if not mermaid_text:
            return {}, []

        node_pattern = re.compile(r'^([A-Za-z0-9_]+)\s*(?:\(\["([^"]+)"\]\)|\["([^"]+)"\])$')
        edge_pattern = re.compile(r'^([A-Za-z0-9_]+)\s*(-->|-.->)(?:\|([^|]+)\|)?\s*([A-Za-z0-9_]+)$')

        nodes: dict[str, dict[str, str]] = {}
        edges: list[dict[str, str | None]] = []

        for raw_line in mermaid_text.splitlines():
            line = raw_line.strip()
            if not line or line.startswith("flowchart"):
                continue
            node_match = node_pattern.match(line)
            if node_match:
                node_id = node_match.group(1)
                nodes[node_id] = {
                    "id": node_id,
                    "label": node_match.group(2) or node_match.group(3) or node_id,
                }
                continue
            edge_match = edge_pattern.match(line)
            if edge_match:
                edges.append(
                    {
                        "from": edge_match.group(1),
                        "kind": "solid" if edge_match.group(2) == "-->" else "dashed",
                        "label": (edge_match.group(3) or "").strip() or None,
                        "to": edge_match.group(4),
                    }
                )

        return nodes, edges

    def _render_lineage_diagram_image(
        self,
        nodes: dict[str, dict[str, str]],
        edges: list[dict[str, str | None]],
        *,
        pil_image_cls,
        image_draw_cls,
        image_font_cls,
    ):
        if pil_image_cls is None or image_draw_cls is None or image_font_cls is None:
            return None

        if not nodes or not edges:
            return None

        indegree: dict[str, int] = {node_id: 0 for node_id in nodes}
        children: dict[str, list[str]] = {node_id: [] for node_id in nodes}
        for edge in edges:
            source = str(edge["from"])
            target = str(edge["to"])
            if source not in nodes or target not in nodes:
                continue
            children[source].append(target)
            indegree[target] = indegree.get(target, 0) + 1

        roots = [node_id for node_id, degree in indegree.items() if degree == 0] or list(nodes.keys())[:1]
        levels: dict[str, int] = {node_id: 0 for node_id in roots}
        queue = list(roots)
        while queue:
            current = queue.pop(0)
            current_level = levels.get(current, 0)
            for child in children.get(current, []):
                next_level = current_level + 1
                if next_level > levels.get(child, -1):
                    levels[child] = next_level
                    queue.append(child)
        for node_id in nodes:
            levels.setdefault(node_id, 0)

        groups: dict[int, list[str]] = {}
        for node_id, level in levels.items():
            groups.setdefault(level, []).append(node_id)
        max_level = max(groups)
        max_group_size = max(len(group) for group in groups.values())

        node_w = 320
        node_h = 72
        gap_x = 120
        gap_y = 42
        margin = 34
        width = margin * 2 + (max_level + 1) * node_w + max_level * gap_x
        height = margin * 2 + max_group_size * node_h + max(0, max_group_size - 1) * gap_y

        image = pil_image_cls.new("RGB", (width, max(height, 280)), "#ffffff")
        draw = image_draw_cls.Draw(image)
        font = image_font_cls.load_default()
        small_font = image_font_cls.load_default()

        positions: dict[str, tuple[int, int, int, int]] = {}
        for level in sorted(groups):
            group = sorted(groups[level], key=lambda node_id: nodes[node_id]["label"])
            total_h = len(group) * node_h + max(0, len(group) - 1) * gap_y
            start_y = margin + max(0, (height - margin * 2 - total_h) // 2)
            x1 = margin + level * (node_w + gap_x)
            for index, node_id in enumerate(group):
                y1 = start_y + index * (node_h + gap_y)
                positions[node_id] = (x1, y1, x1 + node_w, y1 + node_h)

        for edge in edges:
            source = str(edge["from"])
            target = str(edge["to"])
            if source not in positions or target not in positions:
                continue
            sx1, sy1, sx2, sy2 = positions[source]
            tx1, ty1, tx2, ty2 = positions[target]
            start = (sx2, (sy1 + sy2) // 2)
            end = (tx1, (ty1 + ty2) // 2)
            stroke = "#94a3b8" if edge.get("kind") == "dashed" else "#2563eb"
            draw.line([start, end], fill=stroke, width=3)
            arrow = [(end[0], end[1]), (end[0] - 10, end[1] - 6), (end[0] - 10, end[1] + 6)]
            draw.polygon(arrow, fill=stroke)
            label = str(edge.get("label") or "maps to")
            label_box = draw.textbbox((0, 0), label, font=small_font)
            label_w = label_box[2] - label_box[0]
            label_h = label_box[3] - label_box[1]
            label_x = (start[0] + end[0] - label_w) // 2
            label_y = (start[1] + end[1] - label_h) // 2 - 12
            draw.rounded_rectangle(
                [label_x - 8, label_y - 4, label_x + label_w + 8, label_y + label_h + 4],
                radius=10,
                fill="#f8fafc",
                outline="#cbd5e1",
                width=1,
            )
            draw.text((label_x, label_y), label, fill="#334155", font=small_font)

        for node_id, node in nodes.items():
            x1, y1, x2, y2 = positions[node_id]
            fill, border, text = self._diagram_node_palette(node_id)
            draw.rounded_rectangle([x1, y1, x2, y2], radius=16, fill=fill, outline=border, width=2)
            label = self._diagram_display_label(node_id, node["label"])
            text_x = x1 + 12
            text_y = y1 + 12
            for line in self._wrap_diagram_text(label, max_chars=30):
                draw.text((text_x, text_y), line, fill=text, font=font)
                text_y += 14

        buffer = BytesIO()
        buffer.name = "lineage.png"
        image.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer

    @staticmethod
    def _diagram_node_palette(node_id: str) -> tuple[str, str, str]:
        if node_id.startswith("drv_"):
            return "#f3e8ff", "#c084fc", "#6b21a8"
        if node_id.startswith("target_col_") or node_id.startswith("target_tbl_"):
            return "#dcfce7", "#86efac", "#166534"
        return "#e8f4fd", "#93c5fd", "#1d4ed8"

    @staticmethod
    def _diagram_display_label(node_id: str, value: str) -> str:
        segments = [segment for segment in value.split(".") if segment]
        if not segments:
            return value
        if node_id.startswith("target_col_") or node_id.startswith("source_col_"):
            return ".".join(segments[-2:]) if len(segments) >= 2 else segments[-1]
        if node_id.startswith("target_tbl_") or node_id.startswith("source_tbl_") or node_id.startswith("tbl_"):
            return ".".join(segments[-2:]) if len(segments) >= 2 else segments[-1]
        if node_id.startswith("drv_"):
            return ".".join(segments[-3:]) if len(segments) >= 3 else ".".join(segments)
        return ".".join(segments[-3:]) if len(segments) >= 3 else ".".join(segments)

    @staticmethod
    def _wrap_diagram_text(value: str, *, max_chars: int) -> list[str]:
        normalized = value.replace(".", ". ").replace("_", "_ ")
        tokens = normalized.split()
        if not tokens:
            return [value]
        lines: list[str] = []
        current = tokens[0]
        for token in tokens[1:]:
            candidate = f"{current} {token}"
            if len(candidate) <= max_chars:
                current = candidate
            else:
                lines.append(current)
                current = token
        lines.append(current)
        return [line.replace(". ", ".").replace("_ ", "_") for line in lines[:3]]

    @staticmethod
    def _summarize_preview_error(exc: Exception) -> str:
        text = str(exc).strip() or "Preview SQL sample rows could not be loaded."
        return (
            "Workbook download continued without sample preview rows because Snowflake could not execute the preview SQL. "
            f"Reason: {text[:600]}"
        )
