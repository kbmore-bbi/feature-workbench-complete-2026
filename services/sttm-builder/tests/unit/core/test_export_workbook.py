from io import BytesIO

from openpyxl import load_workbook

from app.core.export_workbook import WorkbookExportService
from app.schema.export_workbook import WorkbookExportRequest


class _UnusedSession:
    def sql(self, *_args, **_kwargs):  # pragma: no cover - defensive guard for this unit test
        raise AssertionError("Workbook export preview SQL should not run for this test.")


def _table(name: str) -> dict[str, str]:
    return {"database": "DB", "schema": "SCH", "table": name}


def test_build_workbook_includes_test_case_sheets_and_seed_rows() -> None:
    service = WorkbookExportService(session=_UnusedSession())
    request = WorkbookExportRequest.model_validate(
        {
            "project_name": "Account Billing Migration",
            "created_by": "codex",
            "target_table": _table("ACCOUNT_DETAILS"),
            "source_tables": [_table("DS_ACCOUNTS")],
            "preview_sql": "",
            "test_case_generation": {
                "status": "completed",
                "domain_name": "account_billing",
                "target_layer": "curated",
                "materialization": "incremental",
                "target_model": "account_details",
                "target_table": "DB.SCH.ACCOUNT_DETAILS",
                "test_groups": [
                    {
                        "group": "direct",
                        "target_columns": ["ACCOUNT_ID"],
                    }
                ],
                "seed_files": [
                    {
                        "file_path": "seeds/account_billing/input.csv",
                        "file_type": "SEED_INPUT",
                        "content": "ACCOUNT_ID,ACCOUNT_NAME\n101,Acme Corp\n",
                    }
                ],
                "test_case_document": [
                    {
                        "test_case_id": "TC_001",
                        "group": "direct",
                        "target_attribute": "ACCOUNT_ID",
                        "source_columns": "DS_ACCOUNTS.ACCT_ID",
                        "mapping_rule": "Direct",
                        "test_case_description": "Account id is copied directly from the source.",
                        "test_type": "Positive",
                        "sample_source_input": "ACCT_ID=101",
                        "expected_target_value": "ACCOUNT_ID=101",
                        "confidence": "HIGH",
                    }
                ],
            },
        }
    )

    workbook_bytes = service.build_workbook(request)
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert "Test Case Summary" in workbook.sheetnames
    assert "Test Cases" in workbook.sheetnames
    assert "input" in workbook.sheetnames

    summary_sheet = workbook["Test Case Summary"]
    assert summary_sheet["A2"].value == "Status"
    assert summary_sheet["B2"].value == "completed"
    assert summary_sheet["A8"].value == "Test groups"
    assert summary_sheet["A13"].value == "Group"
    assert summary_sheet["A14"].value == "direct"
    assert summary_sheet["B14"].value == "ACCOUNT_ID"

    document_sheet = workbook["Test Cases"]
    assert document_sheet["A2"].value == "Test Case ID"
    assert document_sheet["A3"].value == "TC_001"
    assert document_sheet["C3"].value == "ACCOUNT_ID"
    assert document_sheet["I3"].value == "ACCOUNT_ID=101"

    seed_sheet = workbook["input"]
    assert seed_sheet["A2"].value == "File Type"
    assert seed_sheet["B2"].value == "SEED_INPUT"
    assert seed_sheet["A4"].value == "ACCOUNT_ID"
    assert seed_sheet["B4"].value == "ACCOUNT_NAME"
    assert seed_sheet["A5"].value == "101"
    assert seed_sheet["B5"].value == "Acme Corp"
